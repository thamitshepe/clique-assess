import openpyxl.styles
import requests
import pandas as pd
import numpy as np
from sklearn.preprocessing import StandardScaler
from sklearn.linear_model import LogisticRegression
import joblib
from sklearn.impute import SimpleImputer

# Function to fetch data for upcoming matches from The Odds API
def fetch_upcoming_matches(api_key):
    url = f'https://api.the-odds-api.com/v4/sports/mma_mixed_martial_arts/odds/?apiKey={api_key}&regions=us&markets=h2h'
    response = requests.get(url)
    data = response.json()
    return data

# Function to preprocess upcoming match data
def preprocess_upcoming_matches(data):
    upcoming_matches = []
    for match in data:
        home_team = match.get('home_team', None)
        away_team = match.get('away_team', None)
        home_odds = None
        away_odds = None
        for bookmaker in match.get('bookmakers', []):
            h2h_market = next((market for market in bookmaker.get('markets', []) if market.get('key') == 'h2h'), None)
            if h2h_market:
                home_odds = next((outcome['price'] for outcome in h2h_market.get('outcomes', []) if outcome['name'] == home_team), None)
                away_odds = next((outcome['price'] for outcome in h2h_market.get('outcomes', []) if outcome['name'] == away_team), None)
                if home_odds is not None and away_odds is not None:
                    break
        # Add a placeholder for the third feature (e.g., match importance)
        third_feature = 0
        upcoming_matches.append({'Home Team': home_team, 'Away Team': away_team, 'Home Odds': home_odds, 'Away Odds': away_odds, 'Third Feature': third_feature})
    return pd.DataFrame(upcoming_matches)

# Function to make predictions using the trained model
def make_predictions(model, upcoming_df):
    # Extract features
    X = upcoming_df[['Home Odds', 'Away Odds']]

    # Check and handle NaN or infinite values
    if np.any(np.isnan(X)) or np.any(np.isinf(X)):
        # Impute NaN values with the median (you can choose a different strategy)
        imputer = SimpleImputer(missing_values=np.nan, strategy='median')
        X = imputer.fit_transform(X)
        # Replace infinities, if any, assuming that infinities are a result of data errors
        X = np.nan_to_num(X, nan=np.median(X), posinf=np.max(X[np.isfinite(X)]), neginf=np.min(X[np.isfinite(X)]))

    # Predict probabilities
    probabilities = model.predict_proba(X)

    probabilities[:, 1] *= 1.1
    probabilities = np.clip(probabilities, 0, 1)  # Ensuring probabilities are between 0 and 1

    # Predict the winner based on adjusted probabilities
    predictions = np.argmax(probabilities, axis=1)

    return predictions, probabilities

# Function to save predictions to an Excel file
def save_predictions_to_excel(predictions, probabilities, upcoming_df, filename='./MMA/MMA_predictions.xlsx'):
    # Map predicted winner codes to actual team names
    upcoming_df['Predicted Winner'] = np.where(predictions == 1, upcoming_df['Home Team'], upcoming_df['Away Team'])
    upcoming_df['Probability (%)'] = np.max(probabilities, axis=1) * 100

    mask = upcoming_df['Probability (%)'] < 65
    upcoming_df['Predicted Winner'] = np.where(mask, upcoming_df['Predicted Winner'] + ' - Risky Bet', upcoming_df['Predicted Winner'])
    
    # Drop the third feature column
    upcoming_df.drop(columns=['Third Feature'], inplace=True)
    
    # Save DataFrame to Excel
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    upcoming_df.to_excel(writer, index=False)
    writer.save()

    # Apply conditional formatting
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    # Define the fill color for risky bets (red)
    red_fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    # Loop through the cells in the 'Predicted Winner' column and apply conditional formatting
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=upcoming_df.columns.get_loc('Predicted Winner') + 1, max_col=upcoming_df.columns.get_loc('Predicted Winner') + 1):
        for cell in row:
            if 'Risky Bet' in cell.value:
                cell.fill = red_fill

    # Save the workbook
    wb.save(filename)

# Main function
def main():
    # Load the trained model
    model = joblib.load('./MMA/trained_model.h5')

    # Fetch upcoming match data
    api_key = '85174e5181a42ac5640cca267ae2ab80'
    upcoming_data = fetch_upcoming_matches(api_key)

    # Preprocess upcoming match data
    upcoming_df = preprocess_upcoming_matches(upcoming_data)

    # Make predictions
    predictions, probabilities = make_predictions(model, upcoming_df)

    # Save predictions to an Excel file
    save_predictions_to_excel(predictions, probabilities, upcoming_df)

if __name__ == "__main__":
    main()